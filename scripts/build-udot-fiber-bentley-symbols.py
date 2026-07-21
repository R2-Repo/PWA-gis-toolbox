"""Consolidate NewSymbols.xlsx Bentley style changelog → bentley-symbols.json."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "js" / "symbology" / "udot-fiber" / "bentley-symbols.json"
DEFAULT_XLSX = Path.home() / "Downloads" / "NewSymbols.xlsx"


def classify(desc: str) -> str:
    d = desc or ""
    if "(Fiber)" in d or "SMF" in d.upper():
        return "fiber"
    if "(Duct)" in d:
        return "duct"
    if any(k in d for k in ("Vault", "JB", "Building", "Splice", "Enclosure", "Cabinet", "RWIS")):
        return "point"
    return "other"


def parse_color(rgb) -> str | None:
    if rgb is None:
        return None
    text = str(rgb).replace("\n", " ").strip()
    nums = re.findall(r"\d+", text)
    if len(nums) < 3:
        return None
    r, g, b = int(nums[0]), int(nums[1]), int(nums[2])
    if r > 255 or g > 255 or b > 255:
        return None
    return f"#{r:02x}{g:02x}{b:02x}"


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        raise SystemExit(f"Workbook not found: {xlsx}")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    by_ms = {}
    for name in wb.sheetnames:
        if not name.startswith("Styles(Update)") or name.endswith("Template"):
            continue
        ws = wb[name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            desc = str(row[0]).replace("\n", " ").strip()
            rgb = row[1] if len(row) > 1 else None
            ms = str(row[2]).replace("\n", " ").strip() if len(row) > 2 and row[2] is not None else ""
            if not ms:
                continue
            by_ms[ms] = {
                "name": ms,
                "description": desc,
                "color": parse_color(rgb),
                "kind": classify(desc),
            }

    payload = {
        "source": "NewSymbols.xlsx consolidated from Styles(Update) sheets",
        "updated": "2023-10-09",
        "symbols": sorted(by_ms.values(), key=lambda x: x["name"].lower()),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"wrote {OUT} ({len(payload['symbols'])} symbols)")


if __name__ == "__main__":
    main()
